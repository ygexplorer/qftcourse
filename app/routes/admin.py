import os
from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app, send_from_directory, jsonify
from flask_login import login_required, current_user
from app.models.chapter import Chapter, Assignment, Submission, Grade
from app.models.announcement import Announcement
from app.models.user import User
from app import db
from functools import wraps

admin_bp = Blueprint('admin', __name__)


def teacher_required(f):
    """教师权限装饰器（教师和助教都可访问）"""
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_staff():
            flash('需要教师或助教权限。', 'error')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated


def teacher_only(f):
    """仅教师权限装饰器（助教不可访问）"""
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_teacher():
            flash('该操作仅限教师。', 'error')
            return redirect(url_for('admin.dashboard'))
        return f(*args, **kwargs)
    return decorated


@admin_bp.route('/')
@teacher_required
def dashboard():
    """教师管理后台首页"""
    total_students = User.query.filter_by(role='student').count()
    total_chapters = Chapter.query.filter_by(is_visible=True).count()
    total_submissions = Submission.query.count()
    # 最新 10 条提交
    recent_submissions = Submission.query.order_by(Submission.submitted_at.desc()).limit(10).all()
    return render_template('admin/dashboard.html',
                           total_students=total_students,
                           total_chapters=total_chapters,
                           total_submissions=total_submissions,
                           recent_submissions=recent_submissions)


# ========== 作业管理 ==========

@admin_bp.route('/submissions')
@teacher_required
def submissions():
    """查看所有提交"""
    chapter_id = request.args.get('chapter', type=int)
    assignment_id = request.args.get('assignment', type=int)

    query = Submission.query
    if assignment_id:
        query = query.filter_by(assignment_id=assignment_id)
    elif chapter_id:
        assignments = Assignment.query.filter_by(chapter_id=chapter_id).all()
        assignment_ids = [a.id for a in assignments]
        query = query.filter(Submission.assignment_id.in_(assignment_ids))

    submissions = query.order_by(Submission.submitted_at.desc()).all()
    chapters = Chapter.query.filter_by(is_visible=True).order_by(Chapter.number).all()

    return render_template('admin/submissions.html',
                           submissions=submissions,
                           chapters=chapters,
                           selected_chapter=chapter_id,
                           selected_assignment=assignment_id)


@admin_bp.route('/score_submission/<int:submission_id>', methods=['POST'])
@teacher_required
def score_submission(submission_id):
    """给提交评分，并同步到成绩表"""
    submission = Submission.query.get_or_404(submission_id)
    score = request.form.get('score', type=float)
    if score is not None:
        if score < 0 or score > 100:
            flash('分数必须在 0-100 之间。', 'error')
            return redirect(request.referrer or url_for('admin.submissions'))
        submission.score = score

        # 同步分数到 Grade 表（按章节汇总）
        chapter_id = submission.assignment.chapter_id
        student_id = submission.student_id
        grade = Grade.query.filter_by(student_id=student_id, chapter_id=chapter_id).first()
        if grade:
            grade.score = score
        else:
            grade = Grade(student_id=student_id, chapter_id=chapter_id, score=score)
            db.session.add(grade)

        db.session.commit()
        flash(f'评分成功：{submission.student.name} 得分 {score}', 'success')
    return redirect(request.referrer or url_for('admin.submissions'))


@admin_bp.route('/download/<int:submission_id>')
@teacher_required
def download(submission_id):
    """教师下载学生作业"""
    submission = Submission.query.get_or_404(submission_id)
    upload_dir = current_app.config['UPLOAD_FOLDER']
    return send_from_directory(
        upload_dir, submission.filepath,
        as_attachment=True, download_name=submission.filename
    )


# ========== 作业发布管理 ==========

@admin_bp.route('/assignments')
@teacher_required
def assignments():
    """管理作业"""
    chapters = Chapter.query.filter_by(is_visible=True).order_by(Chapter.number).all()
    return render_template('admin/assignments.html', chapters=chapters, Submission=Submission)


@admin_bp.route('/assignments/create/<int:chapter_id>', methods=['GET', 'POST'])
@teacher_only
def create_assignment(chapter_id):
    """为某章节创建作业"""
    chapter = Chapter.query.get_or_404(chapter_id)
    if request.method == 'POST':
        title = request.form.get('title', '课后作业')
        description = request.form.get('description', '')
        due_date_str = request.form.get('due_date')

        from datetime import datetime
        due_date = None
        if due_date_str:
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d')

        assignment = Assignment(
            chapter_id=chapter_id,
            title=title,
            description=description,
            due_date=due_date
        )
        db.session.add(assignment)
        db.session.commit()
        flash(f'已为第{chapter.number}章创建作业。', 'success')
        return redirect(url_for('admin.assignments'))

    return render_template('admin/assignment_form.html', chapter=chapter)


@admin_bp.route('/assignments/delete/<int:assignment_id>', methods=['POST'])
@teacher_only
def delete_assignment(assignment_id):
    """删除作业（同时删除相关提交）"""
    assignment = Assignment.query.get_or_404(assignment_id)
    # 删除关联的提交文件
    for sub in assignment.submissions.all():
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], sub.filepath)
        if os.path.exists(filepath):
            os.remove(filepath)
    db.session.delete(assignment)
    db.session.commit()
    flash('作业已删除。', 'info')
    return redirect(url_for('admin.assignments'))


# ========== 成绩管理 ==========

@admin_bp.route('/grades')
@teacher_required
def grades():
    """成绩管理"""
    students = User.query.filter_by(role='student').order_by(User.student_id).all()
    chapters = Chapter.query.filter_by(is_visible=True).order_by(Chapter.number).all()

    # 一次性查询所有成绩，避免 N+1 问题
    all_grades = Grade.query.all()
    grades = {}
    for g in all_grades:
        if g.student_id not in grades:
            grades[g.student_id] = {}
        grades[g.student_id][g.chapter_id] = g

    return render_template('admin/grades.html', students=students, chapters=chapters, grades=grades)


@admin_bp.route('/grades/update', methods=['POST'])
@teacher_required
def update_grade():
    """更新成绩（支持表单和 AJAX JSON 请求）"""
    if request.is_json:
        data = request.get_json()
        student_id = data.get('student_id', type=int) if data else None
        chapter_id = data.get('chapter_id', type=int) if data else None
        score = data.get('score', type=float) if data else None
        comment = data.get('comment', '') if data else ''
    else:
        student_id = request.form.get('student_id', type=int)
        chapter_id = request.form.get('chapter_id', type=int)
        score = request.form.get('score', type=float)
        comment = request.form.get('comment', '')

    if not student_id or not chapter_id:
        if request.is_json:
            return jsonify({'success': False, 'message': '参数错误。'}), 400
        flash('参数错误。', 'error')
        return redirect(url_for('admin.grades'))

    result = _update_grade(student_id, chapter_id, score, comment)

    if request.is_json:
        if result:
            return jsonify({'success': True, 'score': score})
        else:
            return jsonify({'success': False, 'message': '分数必须在 0-100 之间。'}), 400

    if result:
        flash('成绩已更新。', 'success')
    return redirect(url_for('admin.grades'))


def _update_grade(student_id, chapter_id, score, comment=''):
    """内部成绩更新逻辑，返回是否成功"""
    if score is not None and (score < 0 or score > 100):
        return False

    grade = Grade.query.filter_by(student_id=student_id, chapter_id=chapter_id).first()
    if grade:
        grade.score = score
        grade.comment = comment
    else:
        grade = Grade(student_id=student_id, chapter_id=chapter_id, score=score, comment=comment)
        db.session.add(grade)

    # 同步分数到 Submission 表
    assignments = Assignment.query.filter_by(chapter_id=chapter_id).all()
    for assignment in assignments:
        submission = Submission.query.filter_by(
            assignment_id=assignment.id, student_id=student_id
        ).first()
        if submission:
            submission.score = score

    db.session.commit()
    return True


@admin_bp.route('/grades/export')
@teacher_required
def export_grades():
    """导出成绩为 CSV"""
    import csv
    import io
    from flask import make_response

    students = User.query.filter_by(role='student').order_by(User.student_id).all()
    chapters = Chapter.query.filter_by(is_visible=True).order_by(Chapter.number).all()

    # 一次性查询所有成绩，避免 N+1
    all_grades = Grade.query.all()
    grade_map = {}
    for g in all_grades:
        if g.student_id not in grade_map:
            grade_map[g.student_id] = {}
        grade_map[g.student_id][g.chapter_id] = g

    output = io.StringIO()
    writer = csv.writer(output)

    # 表头
    header = ['学号', '姓名', '用户名'] + [f'第{ch.number}章' for ch in chapters] + ['总分', '平均分']
    writer.writerow(header)

    for student in students:
        row = [student.student_id, student.name, student.username]
        total = 0
        count = 0
        student_grades = grade_map.get(student.id, {})
        for ch in chapters:
            g = student_grades.get(ch.id)
            score = g.score if g and g.score is not None else ''
            row.append(score)
            if score != '':
                total += score
                count += 1
        row.append(total)
        row.append(round(total / count, 1) if count > 0 else '')
        writer.writerow(row)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8-sig'
    response.headers['Content-Disposition'] = 'attachment; filename=grades.csv'
    return response


# ========== 学生管理 ==========

@admin_bp.route('/students')
@teacher_required
def students():
    """学生列表"""
    students = User.query.filter_by(role='student').order_by(User.student_id).all()
    return render_template('admin/students.html', students=students)


@admin_bp.route('/students/delete/<int:student_id>', methods=['POST'])
@teacher_only
def delete_student(student_id):
    """删除学生（同时删除其提交和成绩）"""
    student = User.query.get_or_404(student_id)
    if student.role != 'student':
        flash('只能删除学生账号。', 'error')
        return redirect(url_for('admin.students'))

    # 删除提交文件
    for sub in student.submissions.all():
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], sub.filepath)
        if os.path.exists(filepath):
            os.remove(filepath)

    # 删除成绩
    for g in student.grades.all():
        db.session.delete(g)

    db.session.delete(student)
    db.session.commit()
    flash(f'已删除学生 {student.name}。', 'info')
    return redirect(url_for('admin.students'))


@admin_bp.route('/students/import', methods=['GET', 'POST'])
@teacher_only
def import_students():
    """批量导入学生（CSV 或 Excel）"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('请选择文件。', 'error')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('未选择文件。', 'error')
            return redirect(request.url)

        filename = file.filename.lower()
        students_data = []

        # 智能列名映射：支持不同表头命名习惯
        COL_MAP = {
            '学号': 'student_id',
            'student_id': 'student_id',
            'StudentID': 'student_id',
            'studentid': 'student_id',
            'id': 'student_id',
            '编号': 'student_id',
            '姓名': 'name',
            'name': 'name',
            '学生姓名': 'name',
            '名字': 'name',
            '用户名': 'username',
            'username': 'username',
            '账号': 'username',
            '邮箱': 'email',
            'email': 'email',
            '电子邮箱': 'email',
            '电子邮件': 'email',
            '专业': 'major',
            '培养单位': 'department',
            '单位': 'department',
        }

        def map_row(raw_row):
            """将原始行数据映射为标准字段"""
            mapped = {'username': '', 'name': '', 'student_id': '', 'email': ''}
            for key, val in raw_row.items():
                if val is None:
                    continue
                k = str(key).strip()
                v = str(val).strip()
                field = COL_MAP.get(k)
                if field and field in mapped:
                    mapped[field] = v
            return mapped

        try:
            if filename.endswith('.csv'):
                import csv
                import io
                stream = io.TextIOWrapper(file.stream, encoding='utf-8-sig')
                reader = csv.DictReader(stream)
                for row in reader:
                    students_data.append(map_row(row))
            elif filename.endswith('.xlsx'):
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    flash('需要安装 openpyxl 库：pip install openpyxl', 'error')
                    return redirect(request.url)
                wb = load_workbook(file)
                ws = wb.active
                header = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not any(row):
                        continue
                    data = dict(zip(header, row))
                    students_data.append(map_row(data))
            elif filename.endswith('.xls'):
                import xlrd
                import tempfile
                import os
                # xlrd 需要文件路径，保存到临时文件
                tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xls')
                try:
                    file.save(tmp_path)
                    wb = xlrd.open_workbook(tmp_path)
                    ws = wb.sheet_by_index(0)
                    header = [str(ws.cell_value(0, j)) for j in range(ws.ncols)]
                    for i in range(1, ws.nrows):
                        vals = [ws.cell_value(i, j) for j in range(ws.ncols)]
                        if not any(vals):
                            continue
                        data = dict(zip(header, vals))
                        students_data.append(map_row(data))
                finally:
                    os.close(tmp_fd)
                    os.unlink(tmp_path)
            else:
                flash('仅支持 CSV 或 Excel（.xlsx / .xls）格式。', 'error')
                return redirect(request.url)
        except Exception as e:
            flash(f'文件读取失败：{str(e)}', 'error')
            return redirect(request.url)

        if not students_data:
            flash('文件中没有找到学生数据。', 'error')
            return redirect(request.url)

        # 验证并导入
        imported = 0
        skipped = 0
        errors = []
        for data in students_data:
            if not data['name'] or not data['student_id']:
                skipped += 1
                continue
            # 自动生成用户名（如果未提供）
            if not data['username']:
                data['username'] = data['student_id']
            if len(data['username']) < 3:
                errors.append(f"{data['name']}：用户名太短（至少3位）")
                skipped += 1
                continue
            # 检查重复
            if User.query.filter_by(username=data['username']).first():
                errors.append(f"{data['name']}：用户名 {data['username']} 已存在")
                skipped += 1
                continue
            if User.query.filter_by(student_id=data['student_id']).first():
                errors.append(f"{data['name']}：学号 {data['student_id']} 已注册")
                skipped += 1
                continue

            # 设置默认密码为学号后6位
            pwd = data['student_id'][-6:] if len(data['student_id']) >= 6 else data['student_id']
            student = User(
                username=data['username'],
                name=data['name'],
                student_id=data['student_id'],
                email=data['email'] or None,
                role='student'
            )
            student.set_password(pwd)
            db.session.add(student)
            imported += 1

        if imported > 0:
            db.session.commit()

        msg = f'成功导入 {imported} 名学生。'
        if skipped > 0:
            msg += f' 跳过 {skipped} 名。'
        if errors:
            msg += f' 错误：{"; ".join(errors[:5])}'
            if len(errors) > 5:
                msg += f'（还有 {len(errors) - 5} 个）'
        flash(msg, 'success' if not errors else 'warning')
        return redirect(url_for('admin.students'))

    return render_template('admin/import_students.html')


# ========== 课程内容管理 ==========

@admin_bp.route('/chapters')
@teacher_only
def chapters():
    """管理章节"""
    chapters = Chapter.query.order_by(Chapter.number).all()
    return render_template('admin/chapters.html', chapters=chapters)


@admin_bp.route('/chapters/edit/<int:chapter_id>', methods=['GET', 'POST'])
@teacher_only
def edit_chapter(chapter_id):
    """编辑章节内容"""
    chapter = Chapter.query.get_or_404(chapter_id)
    if request.method == 'POST':
        chapter.title = request.form.get('title', chapter.title)
        chapter.content = request.form.get('content', chapter.content)
        chapter.is_visible = 'is_visible' in request.form
        db.session.commit()
        flash('章节已更新。', 'success')
        return redirect(url_for('admin.chapters'))
    return render_template('admin/edit_chapter.html', chapter=chapter)


# ========== 公告管理 ==========

@admin_bp.route('/announcements')
@teacher_required
def announcements():
    """公告列表"""
    all_announcements = Announcement.query.order_by(
        Announcement.is_pinned.desc(),
        Announcement.created_at.desc()
    ).all()
    return render_template('admin/announcements.html', announcements=all_announcements)


@admin_bp.route('/announcements/create', methods=['GET', 'POST'])
@teacher_only
def create_announcement():
    """发布公告"""
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        content = request.form.get('content', '').strip()
        is_pinned = 'is_pinned' in request.form

        if not title or not content:
            flash('标题和内容不能为空。', 'error')
            return redirect(request.url)

        announcement = Announcement(title=title, content=content, is_pinned=is_pinned)
        db.session.add(announcement)
        db.session.commit()
        flash('公告已发布。', 'success')
        return redirect(url_for('admin.announcements'))

    return render_template('admin/announcement_form.html')


@admin_bp.route('/announcements/edit/<int:announcement_id>', methods=['GET', 'POST'])
@teacher_only
def edit_announcement(announcement_id):
    """编辑公告"""
    announcement = Announcement.query.get_or_404(announcement_id)
    if request.method == 'POST':
        announcement.title = request.form.get('title', '').strip()
        announcement.content = request.form.get('content', '').strip()
        announcement.is_pinned = 'is_pinned' in request.form
        db.session.commit()
        flash('公告已更新。', 'success')
        return redirect(url_for('admin.announcements'))

    return render_template('admin/announcement_form.html', announcement=announcement)


@admin_bp.route('/announcements/delete/<int:announcement_id>', methods=['POST'])
@teacher_only
def delete_announcement(announcement_id):
    """删除公告"""
    announcement = Announcement.query.get_or_404(announcement_id)
    db.session.delete(announcement)
    db.session.commit()
    flash('公告已删除。', 'info')
    return redirect(url_for('admin.announcements'))


@admin_bp.route('/announcements/toggle/<int:announcement_id>', methods=['POST'])
@teacher_only
def toggle_announcement(announcement_id):
    """切换公告显示/隐藏"""
    announcement = Announcement.query.get_or_404(announcement_id)
    announcement.is_active = not announcement.is_active
    db.session.commit()
    state = '显示' if announcement.is_active else '隐藏'
    flash(f'公告已{state}。', 'info')
    return redirect(url_for('admin.announcements'))
